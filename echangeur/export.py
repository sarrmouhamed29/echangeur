import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF
from .modele import ResultatFinal


# ──────────────────────────── EXCEL ────────────────────────────

def generer_buffer_excel(r: ResultatFinal) -> bytes:
    wb = openpyxl.Workbook()

    _feuille_resultats(wb, r)
    _feuille_iterations(wb, r)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _style_entete(ws, row, col, texte, largeur_col=None):
    cell = ws.cell(row=row, column=col, value=texte)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if largeur_col:
        ws.column_dimensions[cell.column_letter].width = largeur_col
    return cell


def _bordure(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _feuille_resultats(wb, r: ResultatFinal):
    ws = wb.active
    ws.title = "Résultats"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    date_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells("A1:C1")
    titre = ws["A1"]
    titre.value = f"Dimensionnement Échangeur à Plaques FP22 — {date_str}"
    titre.font = Font(bold=True, size=13, color="FFFFFF")
    titre.fill = PatternFill("solid", fgColor="2E75B6")
    titre.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    def section(row, label):
        ws.merge_cells(f"A{row}:C{row}")
        c = ws[f"A{row}"]
        c.value = label
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="left")

    def ligne(row, param, val, unite=""):
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=3, value=unite)

    # Résultats finaux
    etat = "CONVERGÉ" if r.converge else "NON CONVERGÉ"
    section(2, f"RÉSULTATS FINAUX — {etat} en {r.n_iterations} itération(s)")
    ligne(3, "H global final", round(r.H_global, 2), "W/m²K")
    ligne(4, "Surface d'échange finale", round(r.S_finale, 4), "m²")
    ligne(5, "Nombre de plaques", r.N_plaques, "—")
    ligne(6, "Nombre de canaux par fluide", r.n_canaux, "—")
    ligne(7, f"Perte de charge {r.fluide1.nom}", round(r.delta_P_fluide1, 1), "Pa")
    ligne(8, f"Perte de charge {r.fluide2.nom}", round(r.delta_P_fluide2, 1), "Pa")
    _bordure(ws, 2, 8, 1, 3)

    # Données process
    p = r.process
    section(10, "DONNÉES PROCESS")
    ligne(11, "Puissance à échanger P", p.P, "W")
    ligne(12, "DTLM", p.DTLM, "K")
    ligne(13, "H estimé initial", p.H_est_initial, "W/m²K")
    ligne(14, "Seuil de convergence", p.seuil_convergence, "%")
    ligne(15, "R. encrassement fluide 1", p.R_enc_fluide1, "m²K/W")
    ligne(16, "R. encrassement fluide 2", p.R_enc_fluide2, "m²K/W")
    _bordure(ws, 10, 16, 1, 3)

    # Géométrie
    g = r.geometrie
    section(18, "GÉOMÉTRIE PLAQUE FP22")
    ligne(19, "Surface unitaire plaque s", g.s, "m²")
    ligne(20, "Gap entre plaques b", g.b, "m")
    ligne(21, "Largeur utile l", g.l, "m")
    ligne(22, "Épaisseur paroi e", g.e, "m")
    ligne(23, "Conductivité paroi λ", g.lambda_paroi, "W/mK")
    _bordure(ws, 18, 23, 1, 3)

    # Fluides
    for i, (fl, label) in enumerate([(r.fluide1, "FLUIDE 1"), (r.fluide2, "FLUIDE 2")]):
        base = 25 + i * 8
        section(base, f"{label} — {fl.nom}")
        ligne(base+1, "Débit volumique Q", fl.Q, "m³/s")
        ligne(base+2, "Masse volumique ρ", fl.rho, "kg/m³")
        ligne(base+3, "Viscosité dynamique μ", fl.mu, "Pa·s")
        ligne(base+4, "Chaleur massique cp", fl.cp, "J/kgK")
        ligne(base+5, "Conductivité thermique λ", fl.lambda_, "W/mK")
        _bordure(ws, base, base+5, 1, 3)


def _feuille_iterations(wb, r: ResultatFinal):
    ws = wb.create_sheet("Itérations")

    entetes = [
        "Itér.", "H est. (W/m²K)", "S (m²)", "N plaques", "n canaux",
        f"v {r.fluide1.nom[:8]} (m/s)", "Re1", "Pr1", "Nu1", "h1 (W/m²K)",
        f"v {r.fluide2.nom[:8]} (m/s)", "Re2", "Pr2", "Nu2", "h2 (W/m²K)",
        "H calc. (W/m²K)", "Erreur (%)", "Convergé",
    ]
    largeurs = [7, 16, 10, 10, 10, 14, 10, 10, 10, 14, 14, 10, 10, 10, 14, 16, 12, 10]
    for col, (titre, larg) in enumerate(zip(entetes, largeurs), 1):
        _style_entete(ws, 1, col, titre, larg)

    vert = PatternFill("solid", fgColor="C6EFCE")
    rouge = PatternFill("solid", fgColor="FFC7CE")

    for it in r.iterations:
        row = it.numero + 1
        vals = [
            it.numero, round(it.H_est, 2), round(it.S, 4), it.N, it.n,
            round(it.v1, 4), round(it.Re1, 1), round(it.Pr1, 2), round(it.Nu1, 2), round(it.h1, 1),
            round(it.v2, 4), round(it.Re2, 1), round(it.Pr2, 2), round(it.Nu2, 2), round(it.h2, 1),
            round(it.H_calc, 2), round(it.erreur * 100, 2), "Oui" if it.converge else "Non",
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="center")
        for col in range(1, len(vals) + 1):
            ws.cell(row=row, column=col).fill = vert if it.converge else rouge

    _bordure(ws, 1, len(r.iterations) + 1, 1, len(entetes))


# ──────────────────────────── PDF ────────────────────────────

class _PDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 11)
        self.set_fill_color(46, 117, 182)
        self.set_text_color(255, 255, 255)
        self.cell(0, 10, "Dimensionnement Echangeur a Plaques FP22", fill=True, ln=True, align="C")
        self.set_text_color(0, 0, 0)
        self.ln(2)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(128)
        self.cell(0, 6, f"Page {self.page_no()} - {datetime.now().strftime('%d/%m/%Y')}", align="C")


def generer_buffer_pdf(r: ResultatFinal) -> bytes:
    pdf = _PDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    def section_titre(texte):
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_fill_color(46, 117, 182)
        pdf.set_text_color(255, 255, 255)
        # Remplacer les caractères non-latin1
        t = texte.encode("latin-1", errors="replace").decode("latin-1")
        pdf.cell(0, 7, t, fill=True, ln=True)
        pdf.set_text_color(0, 0, 0)

    def ligne_kv(cle, val, unite=""):
        pdf.set_font("Helvetica", "", 9)
        c = cle.encode("latin-1", errors="replace").decode("latin-1")
        v = str(val)
        u = unite.encode("latin-1", errors="replace").decode("latin-1")
        pdf.cell(95, 6, c, border="B")
        pdf.cell(40, 6, v, border="B", align="R")
        pdf.cell(30, 6, u, border="B", align="C")
        pdf.ln()

    # Résultats finaux
    etat = "CONVERGE" if r.converge else "NON CONVERGE"
    section_titre(f"RESULTATS FINAUX — {etat} en {r.n_iterations} iteration(s)")
    ligne_kv("H global final", round(r.H_global, 2), "W/m2K")
    ligne_kv("Surface d'echange finale", round(r.S_finale, 4), "m2")
    ligne_kv("Nombre de plaques", r.N_plaques, "—")
    ligne_kv("Nombre de canaux par fluide", r.n_canaux, "—")
    nom1 = r.fluide1.nom.encode("latin-1", errors="replace").decode("latin-1")
    nom2 = r.fluide2.nom.encode("latin-1", errors="replace").decode("latin-1")
    ligne_kv(f"Perte de charge {nom1}", round(r.delta_P_fluide1, 1), "Pa")
    ligne_kv(f"Perte de charge {nom2}", round(r.delta_P_fluide2, 1), "Pa")
    pdf.ln(4)

    # Données process
    p = r.process
    section_titre("DONNEES PROCESS")
    ligne_kv("Puissance a echanger P", p.P, "W")
    ligne_kv("DTLM", p.DTLM, "K")
    ligne_kv("H estime initial", p.H_est_initial, "W/m2K")
    ligne_kv("Seuil de convergence", f"{p.seuil_convergence} %", "")
    ligne_kv("R. encrassement fluide 1", p.R_enc_fluide1, "m2K/W")
    ligne_kv("R. encrassement fluide 2", p.R_enc_fluide2, "m2K/W")
    pdf.ln(4)

    # Géométrie
    g = r.geometrie
    section_titre("GEOMETRIE PLAQUE FP22")
    ligne_kv("Surface unitaire plaque s", g.s, "m2")
    ligne_kv("Gap entre plaques b", g.b, "m")
    ligne_kv("Largeur utile l", g.l, "m")
    ligne_kv("Epaisseur paroi e", g.e, "m")
    ligne_kv("Conductivite paroi lambda", g.lambda_paroi, "W/mK")
    pdf.ln(4)

    # Tableau itérations
    section_titre("TABLEAU DES ITERATIONS")
    pdf.set_font("Helvetica", "B", 8)
    cols = ["It.", "H est.", "S(m2)", "N pl.", "h1(W/m2K)", "h2(W/m2K)", "H calc.", "Err.%", "Conv."]
    widths = [10, 20, 18, 14, 24, 24, 20, 16, 14]
    for c, w in zip(cols, widths):
        pdf.cell(w, 6, c, border=1, align="C")
    pdf.ln()
    pdf.set_font("Helvetica", "", 8)
    for it in r.iterations:
        vals = [
            str(it.numero), f"{it.H_est:.1f}", f"{it.S:.3f}", str(it.N),
            f"{it.h1:.0f}", f"{it.h2:.0f}", f"{it.H_calc:.1f}",
            f"{it.erreur*100:.1f}", "Oui" if it.converge else "Non",
        ]
        for v, w in zip(vals, widths):
            pdf.cell(w, 6, v, border=1, align="C")
        pdf.ln()

    buf = io.BytesIO()
    pdf_bytes = pdf.output()
    buf.write(pdf_bytes)
    return buf.getvalue()


# ──────────────────────────── DOCUMENTATION PDF ────────────────────────────

_BLEU      = (46,  117, 182)
_BLEU_CLAIR= (189, 215, 238)
_GRIS      = (242, 242, 242)
_BLANC     = (255, 255, 255)
_NOIR      = (0,   0,   0)
_TEXTE     = (38,  38,  38)


class _DocPDF(FPDF):
    def __init__(self):
        super().__init__()
        self.add_font("Arial",  style="",  fname="/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf")
        self.add_font("Arial",  style="B", fname="/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf")
        self.add_font("Arial",  style="I", fname="/usr/share/fonts/truetype/liberation/LiberationSans-Italic.ttf")
        self.set_auto_page_break(auto=True, margin=18)

    def header(self):
        if self.page_no() == 1:
            return
        self.set_font("Arial", "B", 9)
        self.set_fill_color(*_BLEU)
        self.set_text_color(*_BLANC)
        self.cell(0, 8, "Documentation Technique — Dimensionnement Échangeur à Plaques FP22",
                  fill=True, ln=True, align="C")
        self.set_text_color(*_TEXTE)
        self.ln(2)

    def footer(self):
        self.set_y(-12)
        self.set_font("Arial", "I", 8)
        self.set_text_color(130, 130, 130)
        self.cell(0, 6, f"Page {self.page_no()}   —   FP22 Dimensionnement Tool", align="C")
        self.set_text_color(*_TEXTE)

    # ── Blocs de mise en forme ──────────────────────────────────────

    def titre_section(self, texte, numero=None):
        self.ln(3)
        self.set_fill_color(*_BLEU)
        self.set_text_color(*_BLANC)
        self.set_font("Arial", "B", 11)
        label = f"  {numero}   {texte}" if numero else f"  {texte}"
        self.cell(0, 9, label, fill=True, ln=True)
        self.set_text_color(*_TEXTE)
        self.ln(2)

    def sous_titre(self, texte):
        self.set_font("Arial", "B", 10)
        self.set_text_color(*_BLEU)
        self.cell(0, 7, texte, ln=True)
        self.set_text_color(*_TEXTE)

    def corps(self, texte, indent=0):
        self.set_font("Arial", "", 9)
        self.set_x(self.get_x() + indent)
        self.multi_cell(0, 5.5, texte)

    def formule(self, texte, explication=""):
        self.set_fill_color(*_GRIS)
        self.set_font("Arial", "B", 10)
        self.cell(0, 9, f"    {texte}", fill=True, ln=True)
        if explication:
            self.set_font("Arial", "I", 8.5)
            self.set_text_color(80, 80, 80)
            self.cell(0, 5, f"    {explication}", ln=True)
            self.set_text_color(*_TEXTE)
        self.ln(1)

    def tableau(self, entetes, lignes, largeurs):
        # En-tête
        self.set_fill_color(*_BLEU)
        self.set_text_color(*_BLANC)
        self.set_font("Arial", "B", 8.5)
        for titre, w in zip(entetes, largeurs):
            self.cell(w, 7, f" {titre}", fill=True, border=1)
        self.ln()
        # Lignes alternées
        self.set_text_color(*_TEXTE)
        self.set_font("Arial", "", 8.5)
        for i, ligne in enumerate(lignes):
            self.set_fill_color(*(_GRIS if i % 2 == 0 else _BLANC))
            for val, w in zip(ligne, largeurs):
                self.cell(w, 6.5, f" {val}", fill=True, border=1)
            self.ln()
        self.ln(3)

    def encadre(self, texte, couleur=_BLEU_CLAIR):
        self.set_fill_color(*couleur)
        self.set_font("Arial", "", 9)
        self.multi_cell(0, 6, f"  {texte}  ", fill=True)
        self.ln(1)

    def fleche(self, texte):
        self.set_font("Arial", "", 9)
        self.cell(6, 6, "▶", ln=False)
        self.multi_cell(0, 6, texte)


def generer_documentation_pdf() -> bytes:
    pdf = _DocPDF()

    # ════════════════════════════════════════════════
    # PAGE DE TITRE
    # ════════════════════════════════════════════════
    pdf.add_page()
    pdf.set_fill_color(*_BLEU)
    pdf.rect(0, 0, 210, 297, "F")

    pdf.set_y(60)
    pdf.set_font("Arial", "B", 26)
    pdf.set_text_color(*_BLANC)
    pdf.cell(0, 14, "Dimensionnement", align="C", ln=True)
    pdf.cell(0, 14, "Échangeur à Plaques", align="C", ln=True)
    pdf.set_font("Arial", "B", 32)
    pdf.set_fill_color(*_BLANC)
    pdf.set_text_color(*_BLEU)
    pdf.ln(4)
    pdf.cell(0, 16, "FP22", align="C", fill=False, ln=True)

    pdf.ln(12)
    pdf.set_font("Arial", "", 12)
    pdf.set_text_color(*_BLANC)
    pdf.cell(0, 8, "Documentation Technique", align="C", ln=True)
    pdf.cell(0, 8, "Méthodes de calcul & Corrélations", align="C", ln=True)

    pdf.set_y(230)
    pdf.set_font("Arial", "I", 10)
    pdf.set_text_color(180, 210, 240)
    pdf.cell(0, 7, f"Généré le {datetime.now().strftime('%d/%m/%Y')}", align="C", ln=True)
    pdf.cell(0, 7, "Outil développé en Python / Streamlit", align="C", ln=True)

    # ════════════════════════════════════════════════
    # PAGE 2 — PRÉSENTATION & ALGORITHME
    # ════════════════════════════════════════════════
    pdf.add_page()
    pdf.set_text_color(*_TEXTE)

    pdf.titre_section("Présentation du logiciel", "1")
    pdf.corps(
        "Cet outil calcule le nombre de plaques nécessaire pour un échangeur à plaques de type FP22, "
        "en déterminant le coefficient global d'échange thermique H par une méthode itérative.\n\n"
        "L'échangeur met en contact deux fluides séparés par des plaques métalliques chevronnées. "
        "Le coefficient H dépend du nombre de plaques, qui dépend lui-même de H : "
        "on résout ce problème circulaire par itérations successives jusqu'à convergence."
    )
    pdf.ln(2)

    pdf.titre_section("Algorithme itératif", "2")
    pdf.corps("À chaque itération, les étapes suivantes sont exécutées :")
    pdf.ln(1)

    etapes = [
        ("Estimer H",          "H_est = valeur initiale (défaut 500 W/m²K)"),
        ("Surface d'échange",  "S = P / (H_est × DTLM)"),
        ("Nombre de plaques",  "N = ceil(S / s)  (arrondi au superieur)"),
        ("Nombre de canaux",   "n = (N - 1) / (2 x np)   ou np = nombre de passes"),
        ("Geometrie canal",    "Section A = b x l   |   Diametre hydraulique Dh = 4bl / 2(b+l)"),
        ("Vitesse des fluides","v = Q / (n x A)"),
        ("Re, Pr, Nu, h",      "Calcul pour chaque fluide via la correlation Nu = a * Re^b * Pr^0.33"),
        ("H calcule",          "1/H_calc = 1/h1 + 1/h2 + e/lambda_paroi + R_enc,1 + R_enc,2"),
        ("Convergence ?",      "Erreur = |H_est - H_calc| / H_est  <  seuil (defaut 5 %)"),
        ("Sinon",              "H_est <- H_calc  ->  retour a l'etape 1"),
    ]
    for i, (titre, desc) in enumerate(etapes, 1):
        pdf.set_font("Arial", "B", 9)
        pdf.set_fill_color(*_BLEU_CLAIR)
        pdf.cell(8, 6.5, f"  {i}", fill=True, border=1)
        pdf.set_font("Arial", "B", 9)
        pdf.cell(38, 6.5, f"  {titre}", fill=True, border=1)
        pdf.set_font("Arial", "", 9)
        pdf.set_fill_color(*(_GRIS if i % 2 == 0 else _BLANC))
        pdf.cell(0, 6.5, f"  {desc}", fill=True, border=1, ln=True)
    pdf.ln(4)

    pdf.encadre(
        "Le calcul des pertes de charge est effectué après convergence, "
        "avec les conditions hydrauliques de la dernière itération (Re, v finaux).",
        couleur=(255, 243, 205)
    )

    # ════════════════════════════════════════════════
    # PAGE 3 — FORMULES THERMIQUES
    # ════════════════════════════════════════════════
    pdf.add_page()

    pdf.titre_section("Formules thermiques", "3")

    # ── 3.1 H global
    pdf.sous_titre("3.1  Coefficient global d'échange H")
    pdf.corps(
        "Le coefficient global H traduit la résistance totale au transfert de chaleur "
        "entre les deux fluides. Il est la somme des résistances en série :"
    )
    pdf.ln(1)
    pdf.formule(
        "1/H  =  1/h1  +  1/h2  +  e/λ_paroi  +  R_enc,1  +  R_enc,2",
        "Toutes les résistances en m²K/W — H final en W/m²K"
    )
    pdf.tableau(
        ["Terme", "Signification", "Unité"],
        [
            ["h1, h2",       "Coefficients de convection côté fluide 1 et 2",     "W/m²K"],
            ["e",             "Épaisseur de la paroi (plaque)",                     "m"],
            ["λ_paroi",       "Conductivité thermique de la paroi",                 "W/mK"],
            ["R_enc,1",       "Résistance d'encrassement côté fluide 1",            "m²K/W"],
            ["R_enc,2",       "Résistance d'encrassement côté fluide 2",            "m²K/W"],
        ],
        [30, 110, 30]
    )

    # ── 3.2 Nusselt
    pdf.sous_titre("3.2  Corrélation de Nusselt (plaques chevronnées)")
    pdf.corps(
        "Le nombre de Nusselt Nu quantifie l'intensité de la convection par rapport "
        "à la conduction pure. La corrélation utilisée est :"
    )
    pdf.ln(1)
    pdf.formule(
        "Nu  =  a · Reᵇ · Pr^0.33",
        "Valable pour les deux fluides — a et b dépendent de l'angle β et de Re"
    )
    pdf.corps("Tableau des coefficients a et b :")
    pdf.ln(1)
    pdf.tableau(
        ["Angle β (°)", "Plage de Re",   "a",     "b"],
        [
            ["< 30",    "< 10",          "0,718", "0,349"],
            ["< 30",    "> 10",          "0,348", "0,663"],
            ["45",      "< 10",          "0,718", "0,349"],
            ["45",      "10 – 100",      "0,400", "0,598"],
            ["45",      "> 100",         "0,300", "0,663"],
            ["50",      "< 20",          "0,630", "0,333"],
            ["50",      "20 – 300",      "0,291", "0,591"],
            ["50",      "> 300",         "0,130", "0,732"],
            ["60",      "< 20",          "0,562", "0,326"],
            ["60",      "20 – 400",      "0,306", "0,529"],
            ["60",      "> 400",         "0,108", "0,703"],
            ["> 65",    "< 20",          "0,562", "0,326"],
            ["> 65",    "20 – 500",      "0,331", "0,503"],
            ["> 65",    "> 500",         "0,087", "0,718"],
        ],
        [32, 44, 28, 28]
    )

    # ── 3.3 h convectif
    pdf.sous_titre("3.3  Coefficient de convection h")
    pdf.corps("Une fois Nu connu, h est obtenu par :")
    pdf.ln(1)
    pdf.formule(
        "h  =  Nu · λ_fluide / Dh",
        "λ_fluide = conductivité thermique du fluide [W/mK],  Dh = diamètre hydraulique [m]"
    )

    # ── 3.4 Dh
    pdf.sous_titre("3.4  Diamètre hydraulique du canal")
    pdf.corps(
        "Le canal entre deux plaques est assimilé à un rectangle de largeur l et d'épaisseur b "
        "(avec b << l pour une plaque FP22). Le diamètre hydraulique vaut :"
    )
    pdf.ln(1)
    pdf.formule(
        "Dh  =  4 · b · l  /  2(b + l)  ≈  2b",
        "b = gap entre plaques [m],  l = largeur utile [m]"
    )

    # ════════════════════════════════════════════════
    # PAGE 4 — FORMULES HYDRAULIQUES
    # ════════════════════════════════════════════════
    pdf.add_page()

    pdf.titre_section("Formules hydrauliques — Pertes de charge", "4")

    # ── 4.1 Canaux
    pdf.sous_titre("4.1  Nombre de canaux par passe")
    pdf.corps(
        "Avec N plaques, il y a N−1 canaux au total, répartis équitablement entre les deux fluides "
        "et entre les np passes :"
    )
    pdf.ln(1)
    pdf.formule(
        "n  =  (N − 1)  /  (2 · np)",
        "n = canaux en parallèle par fluide et par passe,  np = nombre de passes"
    )
    pdf.corps(
        "Plus np est grand, moins il y a de canaux en parallèle → la vitesse augmente "
        "→ meilleur échange mais pertes de charge plus élevées."
    )
    pdf.ln(3)

    # ── 4.2 Darcy-Weisbach
    pdf.sous_titre("4.2  Équation de Darcy-Weisbach")
    pdf.corps(
        "La perte de charge dans les canaux est calculée par la formule de Darcy-Weisbach "
        "appliquée à un canal de longueur L = s/l (hauteur de plaque) :"
    )
    pdf.ln(1)
    pdf.formule(
        "ΔP  =  f · (L / Dh) · (ρ · v² / 2)",
        "f = facteur de frottement,  L = longueur de plaque [m],  ρ = masse volumique [kg/m³],  v = vitesse [m/s]"
    )
    pdf.ln(2)

    # ── 4.3 Facteur f
    pdf.sous_titre("4.3  Facteur de frottement f  =  C · Reⁿ")
    pdf.corps(
        "Le facteur de frottement dépend de l'angle de corrugation β et du régime d'écoulement "
        "(nombre de Reynolds Re). Les corrélations utilisées sont :"
    )
    pdf.ln(1)
    pdf.tableau(
        ["Géométrie", "Plage de Re",    "C",      "n",      "Expression"],
        [
            ["Lisse",    "< 2 000",      "24,000", "−1,000", "f = 24 / Re"],
            ["Lisse",    "> 2 000",      "0,079",  "−0,250", "f = 0,079 · Re^−0.25"],
            ["α = 30°",  "40 – 500",     "23,330", "−0,809", "f = 23,33 · Re^−0.809"],
            ["α = 30°",  "500 – 17000",  "0,557",  "−0,211", "f = 0,557 · Re^−0.211"],
            ["α = 60°",  "20 – 140",     "47,450", "−0,680", "f = 47,45 · Re^−0.680"],
            ["α = 60°",  "140 – 4500",   "3,917",  "−0,175", "f = 3,917 · Re^−0.175"],
            ["α = 90°",  "40 – 180",     "63,800", "−0,809", "f = 63,8 · Re^−0.809"],
            ["α = 90°",  "180 – 700",    "4,820",  "−0,312", "f = 4,82 · Re^−0.312"],
        ],
        [22, 30, 18, 18, 52]
    )

    pdf.encadre(
        "Note : pour un angle β donné, le mapping vers la géométrie de référence est :\n"
        "  β < 15° → Lisse   |   15° ≤ β < 45° → α = 30°   |   "
        "45° ≤ β < 75° → α = 60°   |   β ≥ 75° → α = 90°"
    )

    # ════════════════════════════════════════════════
    # PAGE 5 — NOMENCLATURE
    # ════════════════════════════════════════════════
    pdf.add_page()

    pdf.titre_section("Nomenclature", "5")
    pdf.tableau(
        ["Symbole", "Définition",                                      "Unité"],
        [
            ["P",        "Puissance thermique échangée",               "W"],
            ["DTLM",     "Différence de température logarithmique moyenne", "K"],
            ["H",        "Coefficient global d'échange",               "W/m²K"],
            ["h1, h2",  "Coefficients de convection (fluide 1 et 2)", "W/m²K"],
            ["Nu",       "Nombre de Nusselt",                          "—"],
            ["Re",       "Nombre de Reynolds   Re = ρ·v·Dh / μ",      "—"],
            ["Pr",       "Nombre de Prandtl    Pr = μ·cp / λ",        "—"],
            ["s",        "Surface unitaire d'une plaque",              "m²"],
            ["b",        "Gap entre deux plaques (épaisseur canal)",   "m"],
            ["l",        "Largeur utile de la plaque",                 "m"],
            ["e",        "Épaisseur de la paroi",                      "m"],
            ["N",        "Nombre total de plaques",                    "—"],
            ["n",        "Nombre de canaux par fluide et par passe",   "—"],
            ["np",       "Nombre de passes",                           "—"],
            ["A",        "Section d'un canal   A = b × l",            "m²"],
            ["Dh",       "Diamètre hydraulique",                       "m"],
            ["L",        "Longueur de plaque   L = s / l",            "m"],
            ["v",        "Vitesse du fluide dans les canaux",          "m/s"],
            ["Q",        "Débit volumique du fluide",                  "m³/s"],
            ["ρ",        "Masse volumique du fluide",                  "kg/m³"],
            ["μ",        "Viscosité dynamique du fluide",              "Pa·s"],
            ["cp",       "Chaleur massique du fluide",                 "J/kgK"],
            ["λ",        "Conductivité thermique",                     "W/mK"],
            ["β",        "Angle de corrugation des plaques",           "°"],
            ["f",        "Facteur de frottement de Darcy-Weisbach",   "—"],
            ["ΔP",       "Perte de charge dans les canaux",            "Pa"],
            ["R_enc",    "Résistance d'encrassement",                  "m²K/W"],
        ],
        [20, 120, 30]
    )

    buf = io.BytesIO()
    buf.write(pdf.output())
    return buf.getvalue()


def generer_buffer_documentation() -> bytes:
    return generer_documentation_pdf()
